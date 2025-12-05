#!/usr/bin/env python3
"""
NOIZYLAB MASTER UPGRADE & IMPROVE
Full version with Excel + Email integration
Cross-platform (Mac/Windows)
"""

import subprocess
import sys
import os
import sqlite3
import datetime
import platform

# Determine platform
IS_WINDOWS = platform.system() == "Windows"
IS_MAC = platform.system() == "Darwin"

# Database and file paths
if IS_WINDOWS:
    DB_FILE = 'C:\\NOIZYLAB\\data\\upgrade_improve.db'
    EXCEL_FILE = 'C:\\NOIZYLAB\\data\\NOIZYLAB_PROMPTS.xlsx'
else:
    DB_FILE = '/Users/m2ultra/NOIZYLAB/data/upgrade_improve.db'
    EXCEL_FILE = '/Users/m2ultra/NOIZYLAB/data/NOIZYLAB_PROMPTS.xlsx'

# Email credentials from environment variables
EMAIL_USER = os.getenv('EMAIL_USER', 'rp@fishmusicinc.com')
EMAIL_PASS = os.getenv('EMAIL_PASS')
IMAP_SERVER = os.getenv('IMAP_SERVER', 'imap.fishmusicinc.com')
IMAP_PORT = int(os.getenv('IMAP_PORT', '993'))

def upgrade_pip_packages():
    """Upgrade all outdated pip packages"""
    print("Checking for outdated pip packages...")
    try:
        outdated = subprocess.check_output([sys.executable, "-m", "pip", "list", "--outdated"]).decode()
        upgraded = []
        for line in outdated.splitlines()[2:]:
            if line.strip():
                pkg = line.split()[0]
                print(f"Upgrading {pkg}...")
                subprocess.run([sys.executable, "-m", "pip", "install", "--upgrade", pkg])
                upgraded.append(pkg)
        return upgraded
    except Exception as e:
        print(f"Error: {e}")
        return []

def check_system_updates():
    """Check and apply system updates"""
    print("Checking for system updates...")
    updated = []
    if IS_MAC:
        try:
            subprocess.run(["brew", "update"], capture_output=True)
            subprocess.run(["brew", "upgrade"], capture_output=True)
            updated.append("Homebrew packages upgraded")
        except:
            print("Brew not available")
    elif IS_WINDOWS:
        # Windows Update via PowerShell
        try:
            subprocess.run(["powershell", "-Command", "Install-Module PSWindowsUpdate -Force -Scope CurrentUser"], capture_output=True)
            updated.append("Windows Update module installed")
        except:
            pass
    return updated

def check_missing_tools(tools):
    """Check for missing essential tools"""
    print("Checking for missing essential tools...")
    missing = []
    cmd = "where" if IS_WINDOWS else "which"
    for tool in tools:
        if subprocess.call(f"{cmd} {tool}", shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL) != 0:
            missing.append(tool)
    if missing:
        print("Missing tools:", ", ".join(missing))
        # Install with appropriate package manager
        if IS_WINDOWS:
            for tool in missing:
                try:
                    subprocess.run(["choco", "install", tool, "-y"])
                except:
                    print(f"Could not install {tool}")
        elif IS_MAC:
            for tool in missing:
                try:
                    subprocess.run(["brew", "install", tool])
                except:
                    print(f"Could not install {tool}")
    else:
        print("All essential tools are installed.")
    return missing

def improve_configurations():
    """Apply system improvements"""
    print("Applying configuration improvements...")
    improvements = []

    if IS_WINDOWS:
        # Windows Defender
        try:
            subprocess.run(["powershell", "-Command", "Set-MpPreference -DisableRealtimeMonitoring $false"], check=True)
            improvements.append("Windows Defender real-time protection enabled")
        except:
            pass
        # Windows Firewall
        try:
            subprocess.run(["powershell", "-Command", "Set-NetFirewallProfile -Profile Domain,Public,Private -Enabled True"], check=True)
            improvements.append("Windows Firewall enabled")
        except:
            pass
    elif IS_MAC:
        # macOS Firewall
        try:
            subprocess.run(["sudo", "/usr/libexec/ApplicationFirewall/socketfilterfw", "--setglobalstate", "on"], capture_output=True)
            improvements.append("macOS Firewall enabled")
        except:
            pass

    return improvements

def fetch_new_improvements_from_excel():
    """Extract improvements from Excel file"""
    improvements = []
    try:
        from openpyxl import load_workbook
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    description = str(row[0]).strip()
                    if description:
                        improvements.append(("excel", description, datetime.date.today().isoformat()))
            print(f"Extracted {len(improvements)} improvements from Excel")
    except ImportError:
        print("openpyxl not installed - skipping Excel extraction")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    return improvements

def fetch_new_improvements_from_emails():
    """Extract improvements from emails"""
    improvements = []
    if not EMAIL_PASS:
        print("EMAIL_PASS not set - skipping email extraction")
        return improvements

    try:
        import imaplib
        import email as email_lib
        from email.header import decode_header

        mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select('inbox')

        # Search for improvement-related emails
        typ, data = mail.search(None, '(OR SUBJECT "improvement" SUBJECT "upgrade")')
        for num in data[0].split()[-20:]:  # Last 20 matching emails
            typ, msg_data = mail.fetch(num, '(RFC822)')
            msg = email_lib.message_from_bytes(msg_data[0][1])
            subject = msg['subject']
            if subject:
                improvements.append(("email", subject, datetime.date.today().isoformat()))

        mail.logout()
        print(f"Extracted {len(improvements)} improvements from emails")
    except Exception as e:
        print(f"Error reading emails: {e}")
    return improvements

def update_database(upgraded, updated, missing, improvements, excel_improvements, email_improvements):
    """Update SQLite database with all improvements"""
    os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS improvements (
        id INTEGER PRIMARY KEY,
        source TEXT,
        description TEXT,
        date_added TEXT
    )
    ''')

    today = datetime.date.today().isoformat()

    for pkg in upgraded:
        cursor.execute('INSERT INTO improvements (source, description, date_added) VALUES (?, ?, ?)',
                       ("upgrade", f"Upgraded pip package: {pkg}", today))
    for item in updated:
        cursor.execute('INSERT INTO improvements (source, description, date_added) VALUES (?, ?, ?)',
                       ("system", item, today))
    for tool in missing:
        cursor.execute('INSERT INTO improvements (source, description, date_added) VALUES (?, ?, ?)',
                       ("missing", f"Installed missing tool: {tool}", today))
    for imp in improvements:
        cursor.execute('INSERT INTO improvements (source, description, date_added) VALUES (?, ?, ?)',
                       ("improve", imp, today))
    for src, desc, date in excel_improvements + email_improvements:
        cursor.execute('INSERT INTO improvements (source, description, date_added) VALUES (?, ?, ?)',
                       (src, desc, date))

    conn.commit()
    conn.close()
    print("Database updated with all improvements.")

def main():
    print("=" * 60)
    print("NOIZYLAB MASTER UPGRADE & IMPROVE")
    print(f"Platform: {platform.system()} ({platform.machine()})")
    print("=" * 60)

    # 1. Upgrade pip packages
    upgraded = upgrade_pip_packages()

    # 2. System updates
    updated = check_system_updates()

    # 3. Check for missing tools
    essential_tools = ["git", "curl", "python3", "node", "npm"]
    if IS_WINDOWS:
        essential_tools = ["git", "curl", "python", "node", "npm", "code"]
    missing = check_missing_tools(essential_tools)

    # 4. Apply improvements
    improvements = improve_configurations()

    # 5. Extract from Excel
    excel_improvements = fetch_new_improvements_from_excel()

    # 6. Extract from Email
    email_improvements = fetch_new_improvements_from_emails()

    # 7. Update database
    update_database(upgraded, updated, missing, improvements, excel_improvements, email_improvements)

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Packages upgraded: {len(upgraded)}")
    print(f"System updates: {len(updated)}")
    print(f"Missing tools: {len(missing)}")
    print(f"Improvements applied: {len(improvements)}")
    print(f"Excel extractions: {len(excel_improvements)}")
    print(f"Email extractions: {len(email_improvements)}")
    print("\nAll tasks completed!")

if __name__ == "__main__":
    main()
